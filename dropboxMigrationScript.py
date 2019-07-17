# Author: Lukas Geiger
# Date: July 3, 2019
# Project: Python script for acccoi partners which creates Dropbox urls (accroding to directory file system) 
# and inserts them into a Master Excel file. 

from openpyxl import load_workbook
from collections import defaultdict
from django.core.validators import URLValidator
from django.core.exceptions import ValidationError
from pyjarowinkler import distance
from statistics import mean
import os, sys
import urllib.parse
import requests

WORKBOOK = "acccoi_workbook.xlsx"
LOAD_FILENAME_WITH_FILEPATH = "masterFiles.txt"
URL_FINAL_FILE = "endURL_file.txt"
ERROR_FILE = "groupingErrors.xlsx"


"""

    Create File Urls

"""


# Function to make urls according to DropBox guidelines. 
def make():

    f_out = open(URL_FINAL_FILE,"w+")

    file_list = readFilenameWithFilepath()

    # Create dictionaries with Phases urls and Workflow urls 
    end_url_phases_dict, end_url_workflow_dict = createWorkflowPhasesDict(file_list)
        
    # If url is valid then take from respecive dictionary, add href strings, 
    # conduct error checks, and write grouping to f_out.
    writeGroupings(end_url_phases_dict, end_url_workflow_dict, f_out)

    # Additional step which converts filenames to URLs, used for error correction
    createErrorUrls(end_url_workflow_dict, end_url_phases_dict)

    f_out.close()

def readFilenameWithFilepath():
    file_list = []
    f_in = open(LOAD_FILENAME_WITH_FILEPATH,"r")

    # Add file from f_in to list without newline character
    for line in f_in:
        file_list.append(line[:len(line)-1])

    f_in.close()

    return file_list


def createWorkflowPhasesDict(file_list):
    base_url = "https://www.dropbox.com/home"

    end_url_phases_dict = defaultdict(list)
    end_url_workflow_dict = defaultdict(list)

    for filename in file_list: 
        path_to_file_list = [] 
        doc_name = filename.split("/")[-1]
        doc_name_wo_filetype = " ".join(doc_name.split('_')[:-1])
        path_to_file = "/".join(filename.replace(doc_name,"").split("/")[5:])
        first_directory_type = filename.split('/')[-4]

        doc_name = urllib.parse.quote_plus(doc_name)
        
        for word in path_to_file.split("/"):
            path_to_file_list.append(urllib.parse.quote(word))

        path_to_file = "/"+"/".join(path_to_file_list)

        end_URL = base_url+path_to_file+"?preview="+doc_name
        
        if first_directory_type == 'Phases':
            end_url_phases_dict[doc_name_wo_filetype].append(end_URL)
        elif first_directory_type == 'Workflow':
            end_url_workflow_dict[doc_name_wo_filetype].append(end_URL)

    return end_url_phases_dict, end_url_workflow_dict

def writeGroupings(end_url_phases_dict, end_url_workflow_dict, fileToWrite):
    if validate_url(end_url_phases_dict) and validate_url(end_url_workflow_dict):

        validLetters = "ChecklistTemplate123456"
        for key in end_url_phases_dict.keys():
            final_string=""
            for url in end_url_phases_dict[key]:
                file_type = url.split('/')[-1].split('_')[-1][:-5]
                final_string += ("Please access the <a href=\""+
                    url.replace("\n","").replace("/xa0"," ")+"\">"+
                    "".join([char for char in file_type if char in validLetters]).replace("Checkliste", "Checklist")+
                    "</a> for this task here. ")
            fileToWrite.write(final_string+"\n")

        for key in end_url_workflow_dict.keys():
            final_string=""
            for url in end_url_workflow_dict[key]:
                file_type = url.split('/')[-1].split('_')[-1][:-5]
                final_string += ("Please access the <a href=\""+
                    url.replace("\n","").replace("/xa0"," ")+"\">"+
                    "".join([char for char in file_type if char in validLetters]).replace("Checkliste", "Checklist")+
                    "</a> for this task here. ")
            fileToWrite.write(final_string+"\n")


# Validate urls with URLValidator(). Can also check if url exists but that takes more time. 
def validate_url(input_dict):
    val = URLValidator()
    for key in input_dict.keys():
        for url in input_dict[key]:
            try:
                val(url)
                # Test for if url exists:
                # request = requests.get(url)
                # print(url+" -> "+str(request.status_code))
                # if request.status_code != 200:
                #     return False
            except ValidationError:
                print("Not a valid URL")
                return False
    return True

# Parse path to files and filenames from file (in this case 'endURL_file.txt'). 
def getPaths():
    f_in = open(URL_FINAL_FILE,"r")
    path_to_file_list = []
    filename_list = []
    for line in f_in:
        first_split_filepath = urllib.parse.unquote_plus(line).split("\">")[0]
        filename_with_path = "_".join(first_split_filepath.split("_Master")[-1].split("_")[0:-1]).replace("?preview=","")
        filename = filename_with_path.split("/")[-1]
        filepath = "/".join(filename_with_path.split("/")[0:-1])+"/"
        filename_list.append(filename)
        path_to_file_list.append(filepath)

    f_in.close()
    return path_to_file_list, filename_list

# Gets all files in a given directory and sorts them alphabetically (IMPORTANT)
def getFilePathsUnix():
    os.system('find ~/Desktop/acccoi/SCOAP_Master/* -type f -print | sort -V > {}'.format(LOAD_FILENAME_WITH_FILEPATH))

# Gets file paths from argument given
def getFilePaths(): 
    args = sys.argv[1:]
    list_files = [] 

    for filename in args:
        list_files.append(filename)

    return list_files


"""

    Create Excel Document 

"""



# Instert links into excel file according to the correct cell. 
def insertLinks(): 
    similarity_index = []
    bad_matches = defaultdict(list)

    workbook = load_workbook(WORKBOOK)
    worksheet = workbook.active
    f_in = open(URL_FINAL_FILE,"r")
    # f_error = open("groupingErrors.txt", "w+")

    path_list, file_list = getPaths()

    # Create hash table with filename as key and url groupings as values retaining information for phases and workflow sorting
    file_hash_table_phases, file_hash_table_workflow = createTable(f_in, file_list)

    # Get titles from Excel Column D
    title_list = getTitles(worksheet)

    # If path cell - instert nothing. If find exact match btw title and filename - insert url(s). 
    # If find nothing then compare similarity matrix - insert url(s) of max similarity.
    writeToExcel(worksheet, title_list, file_hash_table_phases, bad_matches, similarity_index, 1, 625)
    writeToExcel(worksheet, title_list, file_hash_table_workflow, bad_matches, similarity_index, 626, 1414)


    print("\nAverage similarity index: "+str(mean(similarity_index)))
    print("\nLowest similarity index: "+str(min(similarity_index)))
    print("\nMaximum similarity index: "+str(max(similarity_index)))
    print(bad_matches, len(bad_matches))
    # for key in bad_matches.keys():
    #     f_error.write(str(key)+" -> "+str(bad_matches[key])+"\n")

    # adjustForErrors(bad_matches.keys(), worksheet)

    workbook.save(WORKBOOK)
    f_in.close()
    # f_error.close()

# Create hash table with filename as key and url groupings as values retaining information for phases and workflow sorting
def createTable(f_in, file_list):
    file_hash_table_phases = {}
    file_hash_table_workflow = {}

    count=0
    for line in f_in:
        if "/Phases/" in line: 
            file_hash_table_phases[" ".join(file_list[count].split("_")[:2])+" "+file_list[count].split("_")[-1]] = line[0:-1]
        elif "/Workflow/" in line:
            file_hash_table_workflow[" ".join(file_list[count].split("_")[1:])] = line[0:-1]
        count+=1
    return file_hash_table_phases, file_hash_table_workflow

# If path cell - instert nothing. If find exact match btw title and filename - insert url(s). 
# If find nothing then compare similarity matrix - insert url(s) of max similarity.
def writeToExcel(worksheet, title_list, file_hash_table, bad_matches, similarity_index, start, end):
    for j in range(start,end):
        print("\n"+title_list[j])
        print("----------------------------------------------------------------------")
        if "/Phases" in title_list[j] or "/Workstreams" in title_list[j]:
            print(str(title_list[j])+" -> [TITLE]") 
            worksheet.cell(row=j+1, column=18).value = ""
            continue
        elif title_list[j] in file_hash_table:
            print(str(title_list[j])+" == "+str(file_hash_table[title_list[j]]))
            worksheet.cell(row=j+1, column=18).value = file_hash_table[title_list[j]]
        else:
            similarity, key = findMaxComp(title_list[j], file_hash_table)
            print("Similarity Index: "+str(similarity)+" \n"+str(title_list[j])+" == "+file_hash_table[key])
            similarity_index.append(similarity)
            worksheet.cell(row=j+1, column=18).value = file_hash_table[key]
            if similarity <= .85:
                bad_matches[title_list[j]].append(key)

# Get titles from Excel Column D
def getTitles(worksheet):
    title_list = []
    for i in range(1,1415):
        directory = worksheet.cell(row=i, column=2).value
        prefix = ""
        if len(directory.split("/")) >= 5:
            if "/Phases/" in directory:
                prefix += directory.split("/")[2].split(".")[0]
                prefix += " ["+directory.split("/")[3].split(" ")[-1].lstrip("0")+"] "
            elif "/Workstreams/" in directory:
                prefix += "["+directory.split("/")[3].split(" ")[-1].lstrip("0")+"]"
                prefix += " "+directory.split("/")[-3]+" "
        indv_cell = worksheet.cell(row=i, column=4).value
        title_list.append(prefix+indv_cell.replace("\xa0"," "))

    return title_list

def createErrorUrls(workflow_dict, phases_dict):
    workbook = load_workbook(ERROR_FILE)
    worksheet = workbook.active

    validLetters = "ChecklistTemplate123456"
    for row in range(103, 170): #length of worksheet
        if worksheet.cell(row=row, column=3).value == "no" or worksheet.cell(row=row, column=3).value == "nothing":
            continue
        else:
            final_string=""
            key = " ".join(str(worksheet.cell(row=row, column=3).value).split("_")[0:-1])
            if key in phases_dict.keys():
                for url in phases_dict[key]:
                    file_type = url.split('/')[-1].split('_')[-1][:-5]
                    final_string += ("Please access the <a href=\""+
                        url.replace("\n","").replace("/xa0"," ")+"\">"+
                        "".join([char for char in file_type if char in validLetters]).replace("Checkliste", "Checklist")+
                        "</a> for this task here. ")

            worksheet.cell(row=row, column=3).value = final_string
            print("Corrected: "+key+" -> "+final_string)

    workbook.save(ERROR_FILE)

# Creates excel with all bad_match urls
def createErrorExcel(error_dict):
    workbook = load_workbook(ERROR_FILE)
    worksheet = workbook.active

    row = 2
    i = 0
    for key in error_dict.keys():
        worksheet.cell(row=row, column=1).value = key
        for value in error_dict[key]:
            worksheet.cell(row=row, column=2).value = str(value)
        row+=1

    workbook.save(ERROR_FILE)

"""

Similarity Tests:

    Adjusted for accuracy not speed

"""

# Returns: similarity ratio between two strings
def similar(a, b):
    return distance.get_jaro_distance(a, b, winkler=False, scaling=0.1)

# Method to compare title to all keys in hash table and find the highest similarity.
# Returns: similarity ratio and most similar key. 
def findMaxComp(title, hash_table):
    keys = hash_table.keys()
    max_ratio = 0
    final_key = ""
    for key in keys:
        if key and title:
            similarity_ratio = similar(key, title)
            if similarity_ratio > max_ratio:
                max_ratio = similarity_ratio
                print("Update key: "+key)
                final_key = key
    return max_ratio, final_key




if __name__ == '__main__':
    
    #getFilePathsUnix()
    #make()
    insertLinks()
    #os.system("open {}".format(WORKBOOK))

